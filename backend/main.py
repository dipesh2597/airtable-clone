from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
import socketio
import uvicorn
from typing import Dict, List, Any, Optional
import json
import uuid
from datetime import datetime
import re
import os
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd
from io import StringIO
from validation import validate_value, detect_data_type

# Initialize FastAPI app
app = FastAPI(title="Airtable Clone API", version="1.0.0")

# Initialize Socket.IO
sio = socketio.AsyncServer(
    async_mode='asgi',
    cors_allowed_origins="*"
)

# Create ASGI app
socket_app = socketio.ASGIApp(sio, app)

# Add CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # In production, specify your frontend URL
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# In-memory storage for spreadsheet data and user presence
spreadsheet_data: Dict[str, Any] = {}
active_users: Dict[str, Dict] = {}
user_sessions: Dict[str, str] = {}  # session_id -> user_id
used_colors: set = set()

# --- Excel Persistence Layer ---
EXCEL_PATH = Path("data/spreadsheet.xlsx")

# Ensure Excel file exists and load into memory

def ensure_excel_file():
    if not EXCEL_PATH.exists():
        EXCEL_PATH.parent.mkdir(exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        # Pre-fill with empty values for 26 columns x 100 rows
        for row in range(1, 101):
            for col in range(1, 27):
                ws.cell(row=row, column=col, value="")
        wb.save(EXCEL_PATH)
        print(f"Created new Excel file at {EXCEL_PATH}")


def excel_to_spreadsheet_data():
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    cells = {}
    max_row = max(ws.max_row, 100)
    max_col = max(ws.max_column, 26)
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell_id = f"{get_column_letter(col)}{row}"
            if cell.value not in (None, ""):
                validation_result = validate_value(str(cell.value))
                cells[cell_id] = {
                    'value': validation_result['formatted_value'],
                    'original_value': str(cell.value),
                    'data_type': validation_result['detected_type'],
                    'is_valid': validation_result['is_valid'],
                    'validation_errors': validation_result['errors'],
                    'last_modified_by': 'system',
                    'last_modified_at': datetime.now().isoformat()
                }
    return {
        "cells": cells,
        "columns": max_col,
        "rows": max_row,
        "metadata": {
            "title": "Spreadsheet",
            "created_at": datetime.now().isoformat(),
            "last_modified": datetime.now().isoformat(),
            "filename": "spreadsheet.xlsx"
        }
    }


def write_cell_to_excel(cell_id, value):
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    col_match = re.match(r'^([A-Z]+)', cell_id)
    row_match = re.match(r'[A-Z]+(\d+)$', cell_id)
    if col_match and row_match:
        col_str = col_match.group(1)
        row = int(row_match.group(1))
        col = 0
        for i, char in enumerate(reversed(col_str)):
            col += (ord(char) - ord('A') + 1) * (26 ** i)
        ws.cell(row=row, column=col, value=value)
        wb.save(EXCEL_PATH)
        print(f"Updated {cell_id} in Excel file.")
        return True
    return False

# CSV Import/Export Functions
def csv_to_spreadsheet_data(csv_content: str) -> Dict[str, Any]:
    """Convert CSV content to spreadsheet data format"""
    try:
        # Read CSV content using pandas
        df = pd.read_csv(StringIO(csv_content))
        
        min_rows = 100
        min_cols = 26
        max_row = max(len(df) + 1, min_rows)  # +1 for header row
        max_col = max(len(df.columns), min_cols)
        
        cells = {}
        # Add header row (column names)
        for col_idx in range(1, max_col + 1):
            if col_idx <= len(df.columns):
                col_name = df.columns[col_idx - 1]
            else:
                col_name = ''
            cell_id = f"{get_column_letter(col_idx)}1"
            validation_result = validate_value(str(col_name))
            cells[cell_id] = {
                'value': validation_result['formatted_value'],
                'original_value': str(col_name),
                'data_type': validation_result['detected_type'],
                'is_valid': validation_result['is_valid'],
                'validation_errors': validation_result['errors'],
                'last_modified_by': 'system',
                'last_modified_at': datetime.now().isoformat()
            }
        # Add data rows
        for row_idx in range(max_row - 1):
            for col_idx in range(1, max_col + 1):
                cell_id = f"{get_column_letter(col_idx)}{row_idx + 2}"
                if row_idx < len(df) and col_idx <= len(df.columns):
                    value = df.iloc[row_idx, col_idx - 1]
                else:
                    value = ''
                if pd.notna(value) and value != '':
                    validation_result = validate_value(str(value))
                    cells[cell_id] = {
                        'value': validation_result['formatted_value'],
                        'original_value': str(value),
                        'data_type': validation_result['detected_type'],
                        'is_valid': validation_result['is_valid'],
                        'validation_errors': validation_result['errors'],
                        'last_modified_by': 'system',
                        'last_modified_at': datetime.now().isoformat()
                    }
                else:
                    cells[cell_id] = {
                        'value': '',
                        'original_value': '',
                        'data_type': 'string',
                        'is_valid': True,
                        'validation_errors': [],
                        'last_modified_by': 'system',
                        'last_modified_at': datetime.now().isoformat()
                    }
        
        result = {
            "cells": cells,
            "columns": max_col,
            "rows": max_row,
            "metadata": {
                "title": "Spreadsheet",
                "created_at": datetime.now().isoformat(),
                "last_modified": datetime.now().isoformat(),
                "filename": "spreadsheet.xlsx"
            }
        }
        
        print(f"CSV import result: {len(cells)} cells, {max_col} columns, {max_row} rows")
        print(f"Sample cells: {list(cells.keys())[:10]}")
        
        return result
    except Exception as e:
        print(f"Error converting CSV to spreadsheet data: {e}")
        raise e

def spreadsheet_data_to_csv(data: Dict[str, Any]) -> str:
    """Convert spreadsheet data to CSV format"""
    try:
        # Find the bounds of the data
        if not data['cells']:
            return ""
        
        # Get all cell coordinates
        cell_coords = []
        for cell_id in data['cells'].keys():
            col_match = re.match(r'^([A-Z]+)', cell_id)
            row_match = re.match(r'[A-Z]+(\d+)$', cell_id)
            if col_match and row_match:
                col_str = col_match.group(1)
                row = int(row_match.group(1))
                
                # Convert column letter to number
                col = 0
                for i, char in enumerate(reversed(col_str)):
                    col += (ord(char) - ord('A') + 1) * (26 ** i)
                
                cell_coords.append((row, col, cell_id))
        
        if not cell_coords:
            return ""
        
        # Find bounds
        max_row = max(coord[0] for coord in cell_coords)
        max_col = max(coord[1] for coord in cell_coords)
        
        # Create a 2D array to hold the data
        csv_data = []
        for row in range(1, max_row + 1):
            row_data = []
            for col in range(1, max_col + 1):
                cell_id = f"{get_column_letter(col)}{row}"
                value = data['cells'].get(cell_id, {}).get('value', '')
                row_data.append(str(value))
            csv_data.append(row_data)
        
        # Convert to CSV string
        csv_string = ""
        for row in csv_data:
            csv_string += ",".join(f'"{cell}"' if ',' in cell or '"' in cell else cell for cell in row) + "\n"
        
        return csv_string.rstrip('\n')
    except Exception as e:
        print(f"Error converting spreadsheet data to CSV: {e}")
        raise e

def import_csv_to_excel(csv_content: str):
    """Import CSV data and update the Excel file"""
    try:
        # Convert CSV to spreadsheet data
        new_data = csv_to_spreadsheet_data(csv_content)
        
        # Update the Excel file
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        # Ensure we have a full 26x100 grid
        for row in range(1, 101):  # 100 rows
            for col in range(1, 27):  # 26 columns
                cell_id = f"{get_column_letter(col)}{row}"
                cell_data = new_data['cells'].get(cell_id)
                if cell_data:
                    ws.cell(row=row, column=col, value=cell_data['value'])
                else:
                    # Ensure empty cells are also written
                    ws.cell(row=row, column=col, value="")
        
        # Save the workbook
        wb.save(EXCEL_PATH)
        print(f"Imported CSV data to Excel file: {EXCEL_PATH}")
        
        # Update global spreadsheet data
        global spreadsheet_data
        spreadsheet_data = new_data
        
        return {"success": True, "message": "CSV data imported successfully"}
    except Exception as e:
        print(f"Error importing CSV: {e}")
        return {"success": False, "error": str(e)}

# --- On Startup ---
ensure_excel_file()
spreadsheet_data = excel_to_spreadsheet_data()
print(f"Spreadsheet data: {spreadsheet_data}")

def get_next_color():
    """Get the next available color, excluding blue colors for focus mode"""
    colors = [
        '#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4', '#FFEAA7',
        '#DDA0DD', '#98D8C8', '#F7DC6F', '#BB8FCE', '#85C1E9',
        '#F8C471', '#82E0AA', '#F1948A', '#85C1E9', '#D7BDE2',
        '#F9E79F', '#ABEBC6', '#FAD7A0', '#D5A6BD', '#A9CCE3'
    ]
    
    used_colors_set = {user['color'] for user in active_users.values()}
    available_colors = [color for color in colors if color not in used_colors_set]
    
    if not available_colors:
        # If all colors are used, start over
        return colors[0]
    
    return available_colors[0]

def release_color(color):
    """Release a color back to the pool (no longer needed with new system)"""
    # Colors are now managed dynamically based on active users
    pass

# Initialize default spreadsheet data (26 columns A-Z, 100 rows)
def initialize_spreadsheet():
    global spreadsheet_data
    spreadsheet_data = {
        "cells": {},
        "columns": 26,  # A-Z
        "rows": 100,
        "metadata": {
            "title": "Spreadsheet",
            "created_at": datetime.now().isoformat(),
            "last_modified": datetime.now().isoformat()
        }
    }

# Initialize spreadsheet on startup
initialize_spreadsheet()

# Data persistence functions
def ensure_data_directory():
    """Ensure the data directory exists"""
    data_dir = Path("data")
    data_dir.mkdir(exist_ok=True)
    return data_dir

def save_spreadsheet_to_file(filename: Optional[str] = None):
    """Save current spreadsheet data to a JSON file"""
    data_dir = ensure_data_directory()
    
    if not filename:
        # Generate filename based on current timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"spreadsheet_{timestamp}.json"
    
    filepath = data_dir / filename
    
    # Update metadata before saving
    spreadsheet_data['metadata']['last_modified'] = datetime.now().isoformat()
    spreadsheet_data['metadata']['filename'] = filename
    
    try:
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(spreadsheet_data, f, indent=2, ensure_ascii=False)
        print(f"Spreadsheet saved to {filepath}")
        return {"success": True, "filename": filename, "filepath": str(filepath)}
    except Exception as e:
        print(f"Error saving spreadsheet: {e}")
        return {"success": False, "error": str(e)}

def load_spreadsheet_from_file(filename: str):
    """Load spreadsheet data from a JSON file"""
    data_dir = ensure_data_directory()
    filepath = data_dir / filename
    
    if not filepath.exists():
        raise FileNotFoundError(f"Spreadsheet file not found: {filename}")
    
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            global spreadsheet_data
            spreadsheet_data = json.load(f)
        print(f"Spreadsheet loaded from {filepath}")
        return {"success": True, "filename": filename}
    except Exception as e:
        print(f"Error loading spreadsheet: {e}")
        return {"success": False, "error": str(e)}

def list_saved_spreadsheets():
    """List all saved spreadsheet files"""
    data_dir = ensure_data_directory()
    spreadsheet_files = []
    
    for filepath in data_dir.glob("*.json"):
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                data = json.load(f)
                spreadsheet_files.append({
                    "filename": filepath.name,
                    "title": data.get("metadata", {}).get("title", "Untitled"),
                    "created_at": data.get("metadata", {}).get("created_at", ""),
                    "last_modified": data.get("metadata", {}).get("last_modified", ""),
                    "size": filepath.stat().st_size
                })
        except Exception as e:
            print(f"Error reading file {filepath}: {e}")
    
    # Sort by last modified date (newest first)
    spreadsheet_files.sort(key=lambda x: x["last_modified"], reverse=True)
    return spreadsheet_files

def delete_spreadsheet_file(filename: str):
    """Delete a saved spreadsheet file"""
    data_dir = ensure_data_directory()
    filepath = data_dir / filename
    
    if not filepath.exists():
        raise FileNotFoundError(f"Spreadsheet file not found: {filename}")
    
    try:
        filepath.unlink()
        print(f"Spreadsheet deleted: {filepath}")
        return {"success": True, "filename": filename}
    except Exception as e:
        print(f"Error deleting spreadsheet: {e}")
        return {"success": False, "error": str(e)}

# Socket.IO event handlers
@sio.event
async def connect(sid, environ):
    """Handle user connection"""
    print(f"User connected: {sid}")
    await sio.emit('user_connected', {'sid': sid}, room=sid)

@sio.event
async def disconnect(sid):
    """Handle user disconnection"""
    print(f"User disconnected: {sid}")
    
    if sid in user_sessions:
        user_id = user_sessions[sid]
        if user_id in active_users:
            user_color = active_users[user_id]['color']
            user_name = active_users[user_id]['name']
            release_color(user_color)
            print(f"User {user_name} left, released color {user_color}")
            
            # Send the complete user object for proper frontend handling
            user_data = {
                'user_id': sid,  # Send sid instead of user_id for proper frontend matching
                'name': user_name,
                'color': user_color
            }
            
            print(f"DEBUG: Sending user_left event with user_id={sid} (sid), user_id_internal={user_id}")
            
            del active_users[user_id]
        del user_sessions[sid]
        
        # Broadcast user left to all remaining users
        await sio.emit('user_left', user_data)
        await sio.emit('user_selection_cleared', {'user_id': user_id})  # Send actual user_id, not sid
        print(f"Broadcasted user_left for {user_id}")

@sio.event
async def join_spreadsheet(sid, data):
    global spreadsheet_data
    spreadsheet_data = excel_to_spreadsheet_data()
    user_id = data.get('user_id', str(uuid.uuid4()))
    user_name = data.get('user_name', f'User {user_id[:8]}')
    
    # Check if this user_id already exists and remove old session
    for existing_sid, existing_user_id in user_sessions.items():
        if existing_user_id == user_id and existing_sid != sid:
            print(f"Removing duplicate user session: {existing_user_id}")
            if existing_user_id in active_users:
                old_color = active_users[existing_user_id]['color']
                release_color(old_color)
                del active_users[existing_user_id]
            del user_sessions[existing_sid]
            await sio.emit('user_left', {'user_id': existing_sid})  # Send sid instead of user_id
            await sio.emit('user_selection_cleared', {'user_id': existing_user_id})  # Send actual user_id
    
    user_color = get_next_color()
    
    user_sessions[sid] = user_id
    
    active_users[user_id] = {
        'name': user_name,
        'color': user_color,
        'sid': sid,
        'current_cell': None
    }
    
    print(f"User {user_name} joined with color {user_color}")
    
    await sio.emit('spreadsheet_data', spreadsheet_data, room=sid)
    await sio.emit('active_users', list(active_users.values()), room=sid)
    
    await sio.emit('user_joined', {
        'user_id': sid,  # Send sid instead of user_id for consistency
        'name': user_name,
        'color': user_color
    }, skip_sid=sid)

@sio.event
async def cell_update(sid, data):
    cell_id = data.get('cell_id')
    value = data.get('value')
    user_id = user_sessions.get(sid)
    
    if not cell_id:
        print(f"Invalid cell_update: missing cell_id from sid={sid}")
        return
        
    if not user_id:
        print(f"Invalid cell_update: user_id=None for sid={sid}, user_sessions={user_sessions}")
        return
        
    if user_id not in active_users:
        print(f"Invalid cell_update: user_id={user_id} not in active_users, active_users={list(active_users.keys())}")
        return
        
    user_name = active_users[user_id]['name']
    print(f"Cell update from {user_name} ({user_id}): {cell_id} = '{value}'")
    validation_result = validate_value(value)
    spreadsheet_data['cells'][cell_id] = {
        'value': validation_result['formatted_value'],
        'original_value': value,
        'data_type': validation_result['detected_type'],
        'is_valid': validation_result['is_valid'],
        'validation_errors': validation_result['errors'],
        'last_modified_by': user_id,
        'last_modified_at': datetime.now().isoformat()
    }
    spreadsheet_data['metadata']['last_modified'] = datetime.now().isoformat()
    # Write to Excel file
    write_cell_to_excel(cell_id, value)
    print(f"Validation result: type={validation_result['detected_type']}, valid={validation_result['is_valid']}")
    if not validation_result['is_valid']:
        print(f"Validation errors: {validation_result['errors']}")
    print(f"Broadcasting cell_updated to all other users: {cell_id} = '{validation_result['formatted_value']}'")
    await sio.emit('cell_updated', {
        'cell_id': cell_id,
        'value': validation_result['formatted_value'],
        'original_value': value,
        'data_type': validation_result['detected_type'],
        'is_valid': validation_result['is_valid'],
        'validation_errors': validation_result['errors'],
        'user_id': user_id
    }, skip_sid=sid)

@sio.event
async def cell_selection(sid, data):
    """Handle cell selection updates"""
    cell_id = data.get('cell_id')
    user_id = user_sessions.get(sid)
    
    if not user_id:
        print(f"Invalid cell_selection: user_id=None for sid={sid}, user_sessions={user_sessions}")
        return
    
    if user_id not in active_users:
        print(f"Invalid cell_selection: user_id={user_id} not in active_users, active_users={list(active_users.keys())}")
        return
    
    user_name = active_users[user_id]['name']
    print(f"Cell selection from {user_name} ({user_id}): {cell_id}")
    
    # Parse cell coordinates from cell_id (e.g., "A1" -> row=0, col=0)
    col = 0
    row = 0
    if cell_id:
        col_match = re.match(r'^([A-Z]+)', cell_id)
        row_match = re.match(r'[A-Z]+(\d+)$', cell_id)
        if col_match and row_match:
            col_str = col_match.group(1)
            col = sum((ord(c) - ord('A') + 1) * (26 ** i) for i, c in enumerate(reversed(col_str))) - 1
            row = int(row_match.group(1)) - 1
    
    # Update user's current cell
    active_users[user_id]['current_cell'] = cell_id
    
    print(f"Broadcasting cell_selected to other users: {user_name} selected {cell_id}")
    # Broadcast selection to other users
    await sio.emit('cell_selected', {
        'cell_id': cell_id,
        'user_id': user_id,
        'user_name': active_users[user_id]['name'],
        'user_color': active_users[user_id]['color']
    }, skip_sid=sid)
    
    # Also emit user_selection event for user list updates
    await sio.emit('user_selection', {
        'user_id': user_id,
        'cell': {
            'row': row,
            'col': col
        }
    }, skip_sid=sid)

@sio.event
async def row_operation(sid, data):
    """Handle row insert/delete operations"""
    operation_type = data.get('type')  # 'insert' or 'delete'
    row_index = data.get('index')
    user_id = user_sessions.get(sid)
    
    if user_id not in active_users:
        print(f"Invalid row_operation: user_id={user_id} not in active_users")
        return
    
    user_name = active_users[user_id]['name']
    print(f"Row operation from {user_name} ({user_id}): {operation_type} at index {row_index}")
    
    if operation_type == 'insert':
        # Shift all cells down that are at or after the insert index
        new_cells = {}
        for cell_id, cell_data in spreadsheet_data['cells'].items():
            # Parse cell coordinates
            col_match = re.match(r'^([A-Z]+)', cell_id)
            row_match = re.search(r'(\d+)$', cell_id)
            if col_match and row_match:
                col = col_match.group(1)
                row = int(row_match.group(1)) - 1  # Convert to 0-based
                
                if row >= row_index:
                    # Shift this cell down by one row
                    new_cell_id = f"{col}{row + 2}"  # +2 because we convert back to 1-based and add 1
                    new_cells[new_cell_id] = cell_data
                else:
                    # Keep this cell in the same position
                    new_cells[cell_id] = cell_data
        
        spreadsheet_data['cells'] = new_cells
        
    elif operation_type == 'delete':
        # Remove all cells in the row and shift others up
        new_cells = {}
        for cell_id, cell_data in spreadsheet_data['cells'].items():
            # Parse cell coordinates
            col_match = re.match(r'^([A-Z]+)', cell_id)
            row_match = re.search(r'(\d+)$', cell_id)
            if col_match and row_match:
                col = col_match.group(1)
                row = int(row_match.group(1)) - 1  # Convert to 0-based
                
                if row == row_index:
                    # Delete this cell (don't add to new_cells)
                    continue
                elif row > row_index:
                    # Shift this cell up by one row
                    new_cell_id = f"{col}{row}"  # row is already 0-based, so this becomes row-1+1
                    new_cells[new_cell_id] = cell_data
                else:
                    # Keep this cell in the same position
                    new_cells[cell_id] = cell_data
        
        spreadsheet_data['cells'] = new_cells
    
    # Broadcast operation to all other users
    await sio.emit('row_operation_applied', {
        'type': operation_type,
        'index': row_index,
        'user_id': user_id,
        'cells': spreadsheet_data['cells']
    }, skip_sid=sid)

@sio.event
async def column_operation(sid, data):
    """Handle column insert/delete operations"""
    operation_type = data.get('type')  # 'insert' or 'delete'
    col_index = data.get('index')
    user_id = user_sessions.get(sid)
    
    if user_id not in active_users:
        print(f"Invalid column_operation: user_id={user_id} not in active_users")
        return
    
    user_name = active_users[user_id]['name']
    print(f"Column operation from {user_name} ({user_id}): {operation_type} at index {col_index}")
    
    def col_index_to_letter(index):
        """Convert column index to letter (0=A, 1=B, etc.)"""
        return chr(65 + index)
    
    def letter_to_col_index(letter):
        """Convert column letter to index (A=0, B=1, etc.)"""
        return ord(letter) - 65
    
    if operation_type == 'insert':
        # Shift all cells right that are at or after the insert index
        new_cells = {}
        for cell_id, cell_data in spreadsheet_data['cells'].items():
            # Parse cell coordinates
            col_match = re.match(r'^([A-Z]+)', cell_id)
            row_match = re.search(r'(\d+)$', cell_id)
            if col_match and row_match:
                col_letter = col_match.group(1)
                row = row_match.group(1)
                col = letter_to_col_index(col_letter)
                
                if col >= col_index:
                    # Shift this cell right by one column
                    new_col_letter = col_index_to_letter(col + 1)
                    new_cell_id = f"{new_col_letter}{row}"
                    new_cells[new_cell_id] = cell_data
                else:
                    # Keep this cell in the same position
                    new_cells[cell_id] = cell_data
        
        spreadsheet_data['cells'] = new_cells
        
    elif operation_type == 'delete':
        # Remove all cells in the column and shift others left
        new_cells = {}
        for cell_id, cell_data in spreadsheet_data['cells'].items():
            # Parse cell coordinates
            col_match = re.match(r'^([A-Z]+)', cell_id)
            row_match = re.search(r'(\d+)$', cell_id)
            if col_match and row_match:
                col_letter = col_match.group(1)
                row = row_match.group(1)
                col = letter_to_col_index(col_letter)
                
                if col == col_index:
                    # Delete this cell (don't add to new_cells)
                    continue
                elif col > col_index:
                    # Shift this cell left by one column
                    new_col_letter = col_index_to_letter(col - 1)
                    new_cell_id = f"{new_col_letter}{row}"
                    new_cells[new_cell_id] = cell_data
                else:
                    # Keep this cell in the same position
                    new_cells[cell_id] = cell_data
        
        spreadsheet_data['cells'] = new_cells
    
    # Broadcast operation to all other users
    await sio.emit('column_operation_applied', {
        'type': operation_type,
        'index': col_index,
        'user_id': user_id,
        'cells': spreadsheet_data['cells']
    }, skip_sid=sid)

@sio.event
async def sort_operation(sid, data):
    """Handle column sorting operations"""
    column = data.get('column')
    direction = data.get('direction')  # 'asc', 'desc', or None for clear
    user_id = user_sessions.get(sid)
    
    if user_id not in active_users:
        return
    
    # Handle clear sort operation
    if column is None or direction is None:
        # For clear sort, we don't need to do anything to the data
        # Just broadcast the clear sort state
        await sio.emit('sort_applied', {
            'column': None,
            'direction': None,
            'user_id': user_id,
            'cells': spreadsheet_data['cells']
        })
        return
    
    if direction not in ['asc', 'desc']:
        return
    
    user_name = active_users[user_id]['name']
    print(f"Sort operation from {user_name}: column {column} {direction}")
    
    # Get all rows with data in the sorted column
    rows_with_data = []
    
    for row_index in range(spreadsheet_data['rows']):
        cell_id = f"{chr(65 + column)}{row_index + 1}"
        cell_data = spreadsheet_data['cells'].get(cell_id)
        cell_value = cell_data['value'] if cell_data else ''
        
        if cell_value.strip():
            row_data = {
                'row_index': row_index,
                'sort_value': cell_value,
                'cells': {}
            }
            
            # Collect all cell data for this row
            for col_index in range(spreadsheet_data['columns']):
                row_cell_id = f"{chr(65 + col_index)}{row_index + 1}"
                if row_cell_id in spreadsheet_data['cells']:
                    row_data['cells'][row_cell_id] = spreadsheet_data['cells'][row_cell_id]
            
            rows_with_data.append(row_data)
    
    if not rows_with_data:
        return
    
    # Sort the rows
    def sort_key(row):
        value = row['sort_value'].strip()
        
        # Try to parse as number first
        try:
            return (0, float(value))  # Numbers sort first
        except ValueError:
            return (1, value.lower())  # Then strings (case insensitive)
    
    rows_with_data.sort(key=sort_key, reverse=(direction == 'desc'))
    
    # Create new cell data with sorted rows
    new_cells = {}
    
    # First, copy all cells that are not in rows with data
    for cell_id, cell_data in spreadsheet_data['cells'].items():
        # Parse cell ID to get row index
        match = re.match(r'([A-Z]+)(\d+)', cell_id)
        if match:
            row_index = int(match.group(2)) - 1
            has_data_in_sort_column = any(row['row_index'] == row_index for row in rows_with_data)
            if not has_data_in_sort_column:
                new_cells[cell_id] = cell_data
    
    # Then place the sorted rows
    for new_row_index, row_data in enumerate(rows_with_data):
        for old_cell_id, cell_data in row_data['cells'].items():
            # Parse old cell ID to get column
            match = re.match(r'([A-Z]+)(\d+)', old_cell_id)
            if match:
                col_letter = match.group(1)
                new_cell_id = f"{col_letter}{new_row_index + 1}"
                new_cells[new_cell_id] = cell_data
    
    # Update spreadsheet data
    spreadsheet_data['cells'] = new_cells
    spreadsheet_data['metadata']['last_modified'] = datetime.now().isoformat()
    
    # Broadcast the sorted data to all users
    await sio.emit('sort_applied', {
        'column': column,
        'direction': direction,
        'user_id': user_id,
        'cells': new_cells
    })

@sio.event
async def save_spreadsheet_request(sid, data):
    """Handle save spreadsheet request from client"""
    user_id = user_sessions.get(sid)
    if user_id not in active_users:
        return
    
    filename = data.get('filename')
    result = save_spreadsheet_to_file(filename)
    
    if result['success']:
        # Notify all clients about the save
        await sio.emit('spreadsheet_saved', {
            'filename': result['filename'],
            'user_id': user_id,
            'user_name': active_users[user_id]['name']
        })
    
    # Send result back to the requesting client
    await sio.emit('save_result', result, room=sid)

@sio.event
async def load_spreadsheet_request(sid, data):
    """Handle load spreadsheet request from client"""
    user_id = user_sessions.get(sid)
    if user_id not in active_users:
        return
    
    filename = data.get('filename')
    if not filename:
        await sio.emit('load_result', {'success': False, 'error': 'Filename is required'}, room=sid)
        return
    
    try:
        result = load_spreadsheet_from_file(filename)
        if result['success']:
            # Broadcast the loaded data to all clients
            await sio.emit('spreadsheet_loaded', {
                'data': spreadsheet_data,
                'filename': filename,
                'user_id': user_id,
                'user_name': active_users[user_id]['name']
            })
        
        # Send result back to the requesting client
        await sio.emit('load_result', result, room=sid)
    except Exception as e:
        await sio.emit('load_result', {'success': False, 'error': str(e)}, room=sid)

@sio.event
async def update_title_request(sid, data):
    """Handle spreadsheet title update request"""
    user_id = user_sessions.get(sid)
    if user_id not in active_users:
        return
    
    title = data.get('title', 'Spreadsheet')
    spreadsheet_data['metadata']['title'] = title
    spreadsheet_data['metadata']['last_modified'] = datetime.now().isoformat()
    
    # Broadcast title update to all clients
    await sio.emit('title_updated', {
        'title': title,
        'user_id': user_id,
        'user_name': active_users[user_id]['name']
    })

# HTTP endpoints
@app.get("/")
async def root():
    """Root endpoint"""
    return {"message": "Airtable Clone API", "status": "running"}

@app.get("/api/spreadsheet")
async def get_spreadsheet():
    """Get current spreadsheet data"""
    return spreadsheet_data

@app.get("/api/users")
async def get_active_users():
    """Get list of active users"""
    return list(active_users.values())

@app.post("/api/spreadsheet/reset")
async def reset_spreadsheet():
    """Reset spreadsheet to initial state"""
    initialize_spreadsheet()
    return {"message": "Spreadsheet reset successfully"}

# Data persistence API endpoints
@app.post("/api/spreadsheet/save")
async def save_spreadsheet(data: Dict[str, Any]):
    """Save current spreadsheet data"""
    filename = data.get("filename")
    result = save_spreadsheet_to_file(filename)
    return result

@app.post("/api/spreadsheet/load")
async def load_spreadsheet(data: Dict[str, Any]):
    """Load spreadsheet data from file"""
    filename = data.get("filename")
    if not filename:
        raise HTTPException(status_code=400, detail="Filename is required")
    
    try:
        result = load_spreadsheet_from_file(filename)
        return result
    except FileNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/spreadsheet/list")
async def list_spreadsheets():
    """List all saved spreadsheets"""
    try:
        files = list_saved_spreadsheets()
        return {"files": files}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/spreadsheet/delete/{filename}")
async def delete_spreadsheet(filename: str):
    """Delete a saved spreadsheet"""
    try:
        result = delete_spreadsheet_file(filename)
        return result
    except FileNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/spreadsheet/update-title")
async def update_spreadsheet_title(data: Dict[str, Any]):
    """Update spreadsheet title"""
    title = data.get("title", "Spreadsheet")
    spreadsheet_data['metadata']['title'] = title
    spreadsheet_data['metadata']['last_modified'] = datetime.now().isoformat()
    return {"success": True, "title": title}

# CSV Import/Export API endpoints
@app.post("/api/spreadsheet/import-csv")
async def import_csv(data: Dict[str, Any]):
    """Import CSV data into the spreadsheet"""
    csv_content = data.get("csv_content")
    if not csv_content:
        raise HTTPException(status_code=400, detail="CSV content is required")
    
    try:
        result = import_csv_to_excel(csv_content)
        if result['success']:
            # Ensure the global spreadsheet_data is updated
            global spreadsheet_data
            spreadsheet_data = excel_to_spreadsheet_data()
            
            print(f"Broadcasting updated spreadsheet data: {len(spreadsheet_data.get('cells', {}))} cells")
            
            # Broadcast the updated data to all connected clients
            await sio.emit('spreadsheet_loaded', {
                'data': spreadsheet_data,
                'filename': 'spreadsheet.xlsx',
                'user_id': 'system',
                'user_name': 'System'
            })
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/spreadsheet/export-csv")
async def export_csv():
    """Export spreadsheet data as CSV"""
    try:
        csv_content = spreadsheet_data_to_csv(spreadsheet_data)
        return {
            "success": True,
            "csv_content": csv_content,
            "filename": "spreadsheet.csv"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

if __name__ == "__main__":
    uvicorn.run("main:socket_app", host="0.0.0.0", port=8000, reload=True)